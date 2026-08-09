from datetime import date
from pathlib import Path

from taskmanager.infrastructure.filesystem import TaskFilesystem
from taskmanager.infrastructure.sqlite_repo import SqliteRepository
from taskmanager.services.excel_export import HEADERS, export_tasks_to_excel
from taskmanager.services.settings_service import Settings
from taskmanager.services.task_service import CreateTaskRequest, TaskService


def test_export_excel(tmp_path: Path):
    work = tmp_path / "work"
    work.mkdir()
    settings = Settings(
        work_dir=str(work),
        warning_color="#ff0000",
        highlight_warnings=True,
        warning_lead_days=1,
    )
    repo = SqliteRepository(tmp_path / "e.db")
    service = TaskService(repo, settings, TaskFilesystem(settings))
    project = service.create_project("ProjA")
    service.create_task(
        CreateTaskRequest(
            project_id=project.id,
            number="1",
            description='<a href="https://example.com"><b>Hello</b></a>',
            comment="<i>c</i>",
            date_end=date(2020, 1, 1),
            color="#ffff00",
            create_folder=False,
        )
    )
    dest = tmp_path / "out.xlsx"
    path = export_tasks_to_excel(
        service, dest, project_ids=[project.id], include_hidden=False
    )
    assert path.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "ProjA" in wb.sheetnames
    rows = list(wb["ProjA"].iter_rows(values_only=True))
    assert list(rows[0]) == list(HEADERS)
    # Создана | Приоритет | Номер | Срок | Описание | Комментарий
    assert rows[1][2] == "1"
    assert rows[1][4] == "Hello https://example.com"
    assert rows[1][5] == "c"
    cell = wb["ProjA"].cell(row=2, column=5)
    assert cell.fill.fgColor.rgb in ("00FFFF00", "FFFF00")
    assert cell.font.color.rgb in ("00FF0000", "FF0000")
    repo.close()


def test_export_excel_archive_month_filter(tmp_path: Path):
    work = tmp_path / "work"
    work.mkdir()
    settings = Settings(work_dir=str(work))
    repo = SqliteRepository(tmp_path / "e2.db")
    service = TaskService(repo, settings, TaskFilesystem(settings))
    project = service.create_project("ProjB")
    assert project.id is not None
    active = service.create_task(
        CreateTaskRequest(
            project_id=project.id,
            number="A",
            description="active",
            create_folder=False,
        )
    )
    old = service.create_task(
        CreateTaskRequest(
            project_id=project.id,
            number="OLD",
            description="old archived",
            create_folder=False,
        )
    )
    new = service.create_task(
        CreateTaskRequest(
            project_id=project.id,
            number="NEW",
            description="new archived",
            create_folder=False,
        )
    )
    service.archive_task(old.id)
    service.archive_task(new.id)
    # Force distinct months in DB
    repo._conn.execute(
        "UPDATE tasks SET archive_month = ? WHERE id = ?",
        ("2025_01", old.id),
    )
    repo._conn.execute(
        "UPDATE tasks SET archive_month = ? WHERE id = ?",
        ("2026_08", new.id),
    )
    repo._conn.commit()
    months = repo.list_archive_months(project.id)
    assert months == ["2026_08", "2025_01"]

    dest = tmp_path / "filtered.xlsx"
    path = export_tasks_to_excel(
        service,
        dest,
        project_ids=[project.id],
        include_archived=True,
        archive_months_by_project={project.id: ["2026_08"]},
    )
    from openpyxl import load_workbook

    rows = list(load_workbook(path)["ProjB"].iter_rows(values_only=True))
    numbers = {r[2] for r in rows[1:]}
    assert active.number in numbers
    assert "NEW" in numbers
    assert "OLD" not in numbers
    repo.close()
