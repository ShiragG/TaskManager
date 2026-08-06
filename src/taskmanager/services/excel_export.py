from __future__ import annotations

from datetime import date
from pathlib import Path

from taskmanager.domain import (
    Project,
    Task,
    TaskStatus,
    contrast_foreground,
    html_to_plain_with_urls,
    is_deadline_warning,
)
from taskmanager.services.task_service import ServiceError, TaskService


class ExcelExportError(ServiceError):
    """Excel export failed."""


HEADERS = (
    "Создана",
    "Приоритет",
    "Номер",
    "Срок",
    "Описание",
    "Комментарий",
)


def export_tasks_to_excel(
    service: TaskService,
    dest: Path,
    *,
    project_ids: list[int],
    include_hidden: bool = False,
    include_archived: bool = False,
) -> Path:
    """Export selected projects to an .xlsx workbook (one sheet per project)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ExcelExportError(
            "Для экспорта в Excel установите зависимость openpyxl"
        ) from exc

    if not project_ids:
        raise ExcelExportError("Выберите хотя бы один проект")

    projects_by_id = {p.id: p for p in service.list_projects()}
    selected: list[Project] = []
    for pid in project_ids:
        project = projects_by_id.get(pid)
        if project is None:
            raise ExcelExportError(f"Проект id={pid} не найден")
        selected.append(project)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    settings = service.settings
    today = date.today()

    for project in selected:
        ws = wb.create_sheet(title=_safe_sheet_name(project.name))
        ws.append(list(HEADERS))
        for task in _collect_tasks(
            service,
            project.id,  # type: ignore[arg-type]
            include_hidden=include_hidden,
            include_archived=include_archived,
        ):
            row_values = _task_row(task)
            ws.append(row_values)
            excel_row = ws.max_row
            fill = None
            if task.color:
                fill = PatternFill(
                    fill_type="solid",
                    fgColor=_hex_to_rgb(task.color),
                )
            warn = settings.highlight_warnings and is_deadline_warning(
                task.date_end,
                today=today,
                lead_days=settings.warning_lead_days,
            )
            if warn:
                font_color = _hex_to_rgb(settings.warning_color)
            elif task.color:
                font_color = _hex_to_rgb(contrast_foreground(task.color))
            else:
                font_color = None
            font = Font(color=font_color) if font_color else None
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=excel_row, column=col)
                if fill is not None:
                    cell.fill = fill
                if font is not None:
                    cell.font = font

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(dest)
    except OSError as exc:
        raise ExcelExportError(f"Не удалось сохранить файл: {exc}") from exc
    return dest


def _collect_tasks(
    service: TaskService,
    project_id: int,
    *,
    include_hidden: bool,
    include_archived: bool,
) -> list[Task]:
    tasks: list[Task] = []
    active = service.repo.list_tasks(
        project_id, status=TaskStatus.ACTIVE, only_hidden=None
    )
    for task in active:
        if task.hidden and not include_hidden:
            continue
        tasks.append(task)
    if include_archived:
        archived = service.repo.list_tasks(
            project_id, status=TaskStatus.ARCHIVED, only_hidden=None
        )
        tasks.extend(archived)
    tasks.sort(key=lambda t: t.number.casefold())
    return tasks


def _task_row(task: Task) -> list[object]:
    created = (
        task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else ""
    )
    return [
        created,
        task.priority,
        task.number,
        task.date_end.isoformat() if task.date_end else "",
        html_to_plain_with_urls(task.description),
        html_to_plain_with_urls(task.comment),
    ]


def _hex_to_rgb(hex_color: str) -> str:
    """Return RRGGBB without '#' for openpyxl."""
    raw = hex_color.lstrip("#")
    if len(raw) == 3:
        raw = "".join(c * 2 for c in raw)
    if len(raw) != 6:
        return "000000"
    return raw.upper()


def _safe_sheet_name(name: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in name).strip() or "Project"
    return cleaned[:31]
